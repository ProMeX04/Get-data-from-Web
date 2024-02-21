import subprocess
import os
from tkinter import Tk as tk,StringVar, IntVar,filedialog, ttk
import requests
import pandas as pd
from io import StringIO
# import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from configparser import ConfigParser 
from time import sleep
from datetime import datetime

# Constant Variables
LESSON = (29, 20, 27, 0, 7, 17, 0, 6, 26, 13, 5, 10, 19, 0, 20, 6, 6, 10, 0, 6, 0, 0, 11, 0, 5)
TOTAL = 258
LIST_ADMIN = ["laihieu2714", "Tran_Nhi0306","superadmin","Ninhkfc72"]

class SessionManager:
  """Manages the session for the user."""
  def __init__(self, username, password):
    self.username = username
    self.password = password
    self.session = requests.Session()

  def login(self):
    """Login and retrieve session ID"""
    response = self.session.get("https://laptrinh24h.vn/accounts/login/?next=")
    csrf_token = response.cookies.get("csrftoken")
    payload = {
      "username": self.username,
      "password": self.password,
      "csrfmiddlewaretoken": csrf_token
    }
    self.session.post("https://laptrinh24h.vn/accounts/login/?next=", data=payload)

  def request(self, url):
    """Get response html web"""
    self.login()
    return self.session.get(url)
   
   
class DataFrameManager:
  """Manages the DataFrame operations."""
  @staticmethod
  def get_dataframe(response, current_lesson):
    """Parses the response text and returns a pandas DataFrame with the extracted data."""
    df = pd.read_html(StringIO(response.text))[0]
    df = df.drop(columns=["Organization", "Points"])
    df.iloc[:, 2:TOTAL+1] = df.iloc[:, 2:TOTAL+1].apply(lambda x: x.astype(str).str[0:2])
     
    rows = []
    for i in range(df.shape[0]):
      start_lesson, total = 2, 0
      name = df.iloc[i, 1].split()[0] # Only keep the first part of the name
      if name in LIST_ADMIN:
        continue
      row = [i+1, name]
      for index, lesson in enumerate(LESSON, start=1):
        accepted = (df.iloc[i, start_lesson: start_lesson + lesson] == "10").sum()
        total += accepted
        row.append(f"{accepted}/{lesson}" if accepted or(index < current_lesson and lesson != 0) else "")
        start_lesson += lesson
      row.append(f"{total}/{TOTAL}")
      rows.append(row)
     
    new_df = pd.DataFrame(rows, columns=["Rank", "Name"] + [i + 1 for i in range(1, len(LESSON)+1)] + ["Total"])
    # header of current_lesson is "Current"
    new_df.rename(columns={current_lesson: "Current"}, inplace=True)
    return new_df



class ExcelManager:
  """Manages the Excel operations."""
  @staticmethod
  def format_cell(dataframe, excel_path, current_lesson):
    """Format cell in excel."""
    workbook = Workbook()
    sheet = workbook.active
    for row in dataframe_to_rows(dataframe, index=False, header=True):
      sheet.append(row)
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['A'].width = 5
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    custom_font = Font(name='Calibri', size=10, bold=True, italic=True, color='000000')
    custom_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
   
    def format_cell(cell):
      cell.border = thin_border
      cell.font = custom_font
      cell.alignment = custom_alignment
      if cell.value and len(str(cell.value).split('/')) == 2:
        left, right = cell.value.split('/')
        if int(left) >= int(right):
          cell.fill = fill_green
        elif int(left) >= int(right)//2:
          cell.fill = fill_yellow
        else:
          cell.fill = fill_red
           
    sheet[1][current_lesson].fill = fill_orange
    for row in sheet.iter_rows():
      list(map(format_cell, row))

    workbook.save(excel_path)


class FileManager:
  """Manages the file operations."""
  @staticmethod
  def get_file_path(file_name):
    """Returns the absolute path of the file."""
    # Get the absolute path of the current file
    file_path = os.path.dirname(os.path.abspath(__file__))
    # Define the path of the file
    return os.path.join(file_path, file_name)

from cryptography.fernet import Fernet

class ConfigManager:
  """Manages the configuration."""
  @staticmethod
  def generate_key():
    """Generates a key and save it into a file"""
    key = Fernet.generate_key()
    with open(FileManager.get_file_path("secret.key"), "wb") as key_file:
      key_file.write(key)

  @staticmethod
  def load_key():
    """Loads the key from the current directory named `secret.key`"""
    return open(FileManager.get_file_path("secret.key"), "rb").read()

  @staticmethod
  def encrypt_message(message):
    """
    Encrypts a message
    """
    key = ConfigManager.load_key()
    encoded_message = message.encode()
    f = Fernet(key)
    encrypted_message = f.encrypt(encoded_message)
    return encrypted_message

  @staticmethod
  def decrypt_message(encrypted_message):
    """
    Decrypts an encrypted message
    """
    key = ConfigManager.load_key()
    f = Fernet(key)
    decrypted_message = f.decrypt(encrypted_message)
    return decrypted_message.decode()

  @staticmethod
  def save_config(username, password, current_lesson, path, class_name):
    """Saves the username, password, current_lesson, path, and class_name to a config file."""
    config = ConfigParser()
    config['DEFAULT'] = {'Username': ConfigManager.encrypt_message(username).decode(), 'Password': ConfigManager.encrypt_message(password).decode(), 'CurrentLesson': current_lesson, 'Path': path, 'Class': class_name}
    with open(FileManager.get_file_path('config.ini'), 'w') as configfile:
      config.write(configfile)

  @staticmethod
  def load_config():
    """Loads the username, password, current_lesson, path, and class_name from a config file."""
    config = ConfigParser()
    config.read(FileManager.get_file_path('config.ini'))
    return ConfigManager.decrypt_message(config['DEFAULT']['Username']), ConfigManager.decrypt_message(config['DEFAULT']['Password']), config['DEFAULT']['CurrentLesson'], config['DEFAULT']['Path'], config['DEFAULT']['Class']


class Application(tk):
  """The main application."""
  def __init__(self, *args, **kwargs):
    super().__init__(*args, **kwargs)
    self.progress = ttk.Progressbar(self, orient="horizontal", length=250, mode="determinate")
    self.progress.grid(row=7, columnspan=4)
    self.title("GET DATA")
    self.geometry("295x190")
    self.resizable(False, False)
    self.iconbitmap(FileManager.get_file_path('ico.ico'))
    self.style = ttk.Style(self)
    self.style.theme_use("vista")
    self.username = StringVar()
    self.password = StringVar()
    self.class_name = StringVar()
    self.current_lesson = IntVar()
    self.excel_path = StringVar()
    self.status = StringVar()
    self.create_widgets()
     
  def focus_next(self, event):
    """Move focus to the next widget."""
    event.widget.tk_focusNext().focus()
    return "break"

  def focus_prev(self, event):
    """Move focus to the previous widget."""
    event.widget.tk_focusPrev().focus()
    return "break"

  def create_widgets(self):
    """Creates the widgets."""
    self.bind("<Up>", self.focus_prev)
    self.bind("<Down>", self.focus_next)
    self.bind("<Return>", self.get_data)
    # Bind the Enter key to the get_data method

    self.font = ("cascadia mono", 10)
    ttk.Label(self, text=" Username ", font= self.font, foreground="Green", anchor="w").grid(row=0, column=0, sticky="w")
    ttk.Label(self, text=" Password ", font=self.font, foreground="Green", anchor="w").grid(row=1, column=0, sticky="w")
    ttk.Label(self, text=" Class ", font=self.font, foreground="Green", anchor="w").grid(row=2, column=0, sticky="w")
    ttk.Label(self, text=" Lesson ", font=self.font, foreground="Green", anchor="w").grid(row=3, column=0, sticky="w")
    ttk.Label(self, text=" Path ", font=self.font, foreground="Green", anchor="w").grid(row=4, column=0, sticky="w")
    ttk.Label(self, textvariable=self.status,font=self.font, foreground="RED").grid(row=6, columnspan=3)
    ttk.Entry(self, textvariable=self.username, font= self.font).grid(row=0, column=1)
    ttk.Entry(self, textvariable=self.password, show="*", font= self.font).grid(row=1, column=1)
    ttk.Entry(self, textvariable=self.class_name, font= self.font).grid(row=2, column=1)
    ttk.Entry(self, textvariable=self.current_lesson, font= self.font).grid(row=3, column=1)
    ttk.Entry(self, textvariable=self.excel_path, font= self.font).grid(row=4, column=1)
    ttk.Button(self, text="...", command=self.browse_folder,width=5).grid(row=4, column=2)
    ttk.Button(self, text="Get", command=self.get_data).grid(row=5, column=1)

  def browse_folder(self):
    """Browse for Excel path."""
    path = filedialog.askdirectory()
    self.excel_path.set(path)

  def get_data(self, event=None):
    """Button."""  
    self.progress["value"] = 0
    self.progress["maximum"] = 100
    self.update_idletasks()
    for i in range(1, 101):
      self.progress["value"] = i
      self.update_idletasks()
      sleep(0.001)
     
    username = self.username.get()
    password = self.password.get()
    current_lesson = self.current_lesson.get()
    path = self.excel_path.get()
    class_name = self.class_name.get()
     
    self.file_name = "/{}{}.xlsx".format(class_name.upper(),datetime.now().strftime("(%d-%m-%Y_%H)"))
    ConfigManager.save_config(username, password, current_lesson, path, class_name)
    session_manager = SessionManager(username, password)
    response = session_manager.request("https://laptrinh24h.vn/contest/{}/ranking/".format(class_name.lower()))
    if response.status_code == 200:
      self.status.set("Login Success")
      dataframe = DataFrameManager.get_dataframe(response, current_lesson)
      ExcelManager.format_cell(dataframe, path + self.file_name, current_lesson)
      subprocess.Popen(["start", path + self.file_name], shell=True)
    else:
      self.status.set("Username or Password is wrong")


def main():
  """main"""
  # ConfigManager.generate_key()
  # ConfigManager.save_config("laihieu2714", "laihieu2714", "1", "D:/data", "1")
   
  app = Application()
  username, password, current_lesson, path, class_name = ConfigManager.load_config() # Load the username, password, current_lesson, path, and class_name
  app.username.set(username)
  app.password.set(password)
  app.current_lesson.set(current_lesson)
  app.excel_path.set(path)
  app.class_name.set(class_name)
  app.mainloop()



if __name__ == "__main__":
  main()
