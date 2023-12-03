import subprocess
import requests
import pandas as pd
from io import StringIO
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import subprocess

"""User Data"""
# credentials
USERNAME = "laihieu2714"
PASSWORD = "iiviizedz"

#create list lesson 
LESSON = (29, 20, 27, 0, 7, 17, 0, 6, 26, 13, 5, 10, 19, 0, 20, 6, 6, 10, 0, 6, 0, 0, 11, 0, 5)
TOTAL = 258
CURRENT_LESSON = 26

# class
CLASS = input("Enter class: ")

# excel path
EXCEL_PATH = "BXH.xlsx"

# list admin
LIST_ADMIN = ["laihieu2714", "Tran_Nhi0306","superadmin", "Ninhkfc72", "Nguyen_Thi_Lua_FullHouse"]


def login(username, password) -> str:
    """Login and retrieve session ID"""
    session = requests.Session()
    
    # Get CSRF token
    response = session.get("https://laptrinh24h.vn/accounts/login/?next=" )
    csrf_token = response.cookies.get("csrftoken")

    payload = {
        "username": username,
        "password": password,
        "csrfmiddlewaretoken": csrf_token
    }
    session.post("https://laptrinh24h.vn/accounts/login/?next=" , data=payload)
    
    session_id = session.cookies.get("sessionid")
    return session_id


def request(url) -> requests.Response:
    """Get response html web"""
    session = requests.Session()
    session.cookies['sessionid'] = login(USERNAME, PASSWORD)
    response = session.get(url)
    return response


def get_data_frame(response) -> pd.DataFrame:
    """
    Parses the response text and returns a pandas DataFrame with the extracted data.

    Parameters:
    response (requests.Response): The response object containing the HTML data.

    Returns:
    pandas.DataFrame: The DataFrame containing the parsed data.
    """
    df = pd.read_html(StringIO(response.text))[0]
    # Rest of the code...
    
    df = df.drop(columns=["Organization"])
    df = df.drop(columns=["Points"])

    # Delete column Organization
    for i in range(1, 259):
        head = str(i) + "  10"
        df[head] = df[head].astype(str).str[0:2]

    new_df = pd.DataFrame()
    so_hoc_vien = df.shape[0]

    for i in range(so_hoc_vien):
        start_lesson, total = 2, 0
        row = [i+1, df.iloc[i, 1].split()[0]]

        if row[1] in LIST_ADMIN:
            continue

        for index, lesson in enumerate(LESSON, start=1):
            accepted = (
                df.iloc[i, start_lesson: start_lesson + lesson] == "10").sum()
            total += accepted
            if accepted:
                row.append(f"{accepted}/{lesson}")
            else:
                if index < CURRENT_LESSON and lesson != 0:
                    row.append(f"{accepted}/{lesson}")
                else:
                    row.append("")

            start_lesson += lesson
        row.append(f"{total}/{TOTAL}")
        new_df = pd.concat(
            [new_df, pd.DataFrame([row])], ignore_index=True)

    new_df = new_df.rename(columns={0: "Rank", 1: "Name", 27: "Total"})

    return new_df


def format_cell(df) -> None:
    """
    Format cell in excel.

    Args:
        df (pandas.DataFrame): The DataFrame containing the data to be formatted.

    Returns:
        None
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['A'].width = 5

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    custom_font = Font(name='Calibri', size=10, bold=True,
                       italic=True, color='000000')
    custom_alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=False)

    fill_green = PatternFill(start_color='00FF00',
                             end_color='00FF00', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00',
                              end_color='FFFF00', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000',
                           end_color='FF0000', fill_type='solid')

    for row in sheet.iter_rows():
        for cell in row:
            # Rest of the code...
            cell.border = thin_border
            cell.font = custom_font
            cell.alignment = custom_alignment
            if cell.value:
                if len(str(cell.value).split('/')) == 2:
                    # Rest of the code...
                    pass
                    left, right = cell.value.split('/')
                    if int(left) >= int(right):
                        cell.fill = fill_green
                    elif int(left) >= int(right)//2:
                        cell.fill = fill_yellow
                    else:
                        cell.fill = fill_red

    workbook.save(EXCEL_PATH)


def main():
    """Main function"""
    res = request("https://laptrinh24h.vn/contest/cpp{}/ranking/".format(CLASS))
    if res.status_code == 200:
        print("Login success")
        dataframe = get_data_frame(res)
        format_cell(dataframe)
        subprocess.Popen(["start", EXCEL_PATH], shell=True)

    else:
        print("Password or username is wrong")


if __name__ == "__main__":
    main()