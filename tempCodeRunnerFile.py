""" main """
import subprocess
import requests
import pandas as pd
import source
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


def request(url):
    """Get response html web"""
    session = requests.Session()
    session.cookies['sessionid'] = source.SESSION_ID
    response = session.get(url)
    return response


def get_data_frame(response):
    """_summary_
    """
    df = pd.read_html(response.text)[0]
    # Read Table from web
    df = df.drop(columns=["Organization"])
    df = df.drop(columns=["Points"])

    # Delete column Organization
    for i in range(1, 259):
        head = str(i) + "  10"
        df[head] = df[head].astype(str).str[0:2]

    # print(df)
    new_df = pd.DataFrame()
    so_hoc_vien = df.shape[0]

    # print(df.head().columns)
    for i in range(so_hoc_vien):
        start_lesson = 2
        row = [i+1, df.iloc[i, 1].split()[0]]
        if row[1] in source.LIST_ADMIN:
            continue

        total = 0
        for lesson in source.LESSON:
            accepted = (
                df.iloc[i, start_lesson: start_lesson + lesson] == "10").sum()
            total += accepted
            if accepted:
                row.append(f"{accepted}/{lesson}")
            else:
                row.append("")
            start_lesson += lesson
        row.append(f"{total}/{source.TOTAL}")
        new_df = pd.concat(
            [new_df, pd.DataFrame([row])], ignore_index=True)

    new_df = new_df.rename(columns={0: "Rank", 1: "Name", 27: "Total"})

    # for NAME in new_df["Name"]:
    #     if NAME in source.LIST_ADMIN:
    #         new_df = new_df.drop(new_df[new_df["Name"] == NAME].index)
    return new_df


def format_cell(df):
    """format cell in excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    # print(sheet)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    custom_font = Font(name='Calibri', size=10, bold=True,
                       italic=True, color='000000')
    custom_alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=False)

    # rule = CellIsRule(
    #     formula=[
    #         'INT(LEFT(A1, FIND("/", A1) - 1)) >= INT(RIGHT(A1, LEN(A1) - FIND("/", A1))'],
    #     stopIfTrue=False,
    #     fill=PatternFill(start_color='00FF00',
    #                      end_color='00FF00', fill_type='solid')
    # )
    fill_green = PatternFill(start_color='00FF00',
                       end_color='00FF00', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00',
                       end_color='FFFF00', fill_type='solid')
    fill_red = PatternFill(start_color='FF0000',
                       end_color='FF0000', fill_type='solid')


    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.font = custom_font
            cell.alignment = custom_alignment
            if cell.value:
                if len(str(cell.value).split('/')) == 2:
                    left, right = cell.value.split('/')
                    if int(left) >= int(right):
                        cell.fill = fill_green
                    elif int(left) >= int(right)//2:
                        cell.fill = fill_yellow
                    else:
                        cell.fill = fill_red

    workbook.save("BXH.xlsx")

if __name__ == "__main__":
    res = request(source.URL)
    if res.status_code == 200:
        print("Success")
        dataframe = get_data_frame(res)
        format_cell(dataframe)
        subprocess.Popen(["start", source.EXCEL], shell=True)

        # dataframe_to_excel(dataframe)
    else:
        print("Can't connect")
